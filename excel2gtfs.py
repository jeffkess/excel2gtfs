""""------------------------------------------------------------------
Excel2GTFS v0.2.1
(c) Jeff Kessler, 2024-03-03-0920

0.0.1  Initial Commit
0.0.2  Schedule data processing
0.0.3  Support for calendar dates and overrides
0.0.4  GTFS specification conformity adjustments
0.0.5  Post-midnight trip support & config sheets
0.0.6  Adds feed info support and attribution
0.0.7  Converts to a function for variable filename operation
0.0.8  Suppresses openpyxl warnings
0.1.0  Adds support for trip_short_names; trip error handling
0.1.1  Expanded post-midnight support depending on field type
0.1.2  Adds block_id support, option to skip spreadsheets
0.2.0  Supports routes serving the same stop twice and
       simultaneous departures from the same origin (in trip_ids)
0.2.1  Prevents exporting of empty files
------------------------------------------------------------------"""

import openpyxl
import csv
import os
import datetime
import sys
import warnings

# Suppress openpyxl warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


def excel2gtfs(filename=None):
    """Function to convert an excel template to GTFS"""

    # Select Workbook and Load
    wb = openpyxl.load_workbook(filename if filename else "excel2gtfsTemplate.xlsm", data_only=True)

    # Identify applicable sheets
    config_sheets = {"Agency", "Routes", "Stops", "Fare Rules", "Fare Attributes", "Shapes", "Calendar", "Calendar Dates", "Feed Info"}
    skip_sheets = {"Settings & Checks"}
    services = set(wb.sheetnames) - config_sheets - skip_sheets
    services = [row for row in services if row[:5] != "SKIP-"]
    config_sheets = config_sheets.intersection(wb.sheetnames) - skip_sheets

    # Create Output Directory
    fp = "Excel2GTFS Output Created " + datetime.datetime.now().strftime("%Y-%m-%d-%H%M%S")
    os.makedirs(fp)


    # ------------------------------------------------------------------------------
    # Process and Write Configuration Sheets
    # ------------------------------------------------------------------------------

    for sheet_name in config_sheets:

        # Parse Dates in GTFS Format
        data = wb[sheet_name].values
        data = [[(val.strftime("%Y%m%d") if type(val)==datetime.datetime else val) for val in row] for row in data]

        # Process Calendar Overrides
        if sheet_name == "Calendar Dates" and data[1:]:

            override_dates = {}
            calendar_dates = []

            # Process Override Entries and Extract Type 3s
            for row in data[1:]:

                # Skip blank rows
                if not any(row):
                    continue

                # Convert to List of Dicts
                row = {data[0][index]: val for index, val in enumerate(row)}

                # Extract Type 3s or Add Generic Calendar Dates
                if str(row["exception_type"])=="3":
                    override_dates[row["date"]].append(row["service_id"]) if row["date"] in override_dates else override_dates.update({row["date"]: [row["service_id"]]})
                else:
                    calendar_dates.append(row)

            # Covert Override Entries to Type 1/2s
            for date, svcs in override_dates.items():
                [calendar_dates.append({"service_id": svc, "date": date, "exception_type": ("1" if svc in svcs else "2")}) for svc in services]

            # Covert Calendar Dates back to List vs Dict
            data = [list(calendar_dates[0])] + [[row[key] for key in list(calendar_dates[0])] for row in calendar_dates[1:]] if calendar_dates else []

        # Append excel2gtfs Attribution
        if sheet_name == "Feed Info" and data[1:]:
            for row in data[1:]:
                row[data[0].index("feed_publisher_name")] += " (Created via the excel2gtfs tool)"

        # Save GTFS Configuration File if data
        if len(data) > 1:
            with open(f'{fp}/{sheet_name.lower().replace(" ", "_")}.txt', "w") as file:
                writer = csv.writer(file)
                writer.writerows(data)


    # ------------------------------------------------------------------------------
    # Process Schedule Data
    # ------------------------------------------------------------------------------

    # Initialize Services
    special_keys = ["route_id", "direction_id", "trip_short_name", "shape_id", "headsign", "wheelchair_accessible", "bikes_allowed", "block_id", "Then Every", "Until"]
    gtfs_entries = {"trips": [], "stop_times": [], "frequencies": []}

    # Process Schedule Sheets
    for service in services:

        # Load trips on the given service
        svc_trips = list(wb[service].values)
        svc_trips = [[(svc_trips[1][index], val) for index, val in enumerate(row)] for row in svc_trips[2:] if any(row)]

        # Track trip_ids to prevent duplicates where two depart at the same time
        trip_ids = []

        for trip in svc_trips:

            # Convert times to list and data to dictionary (per special_keys)
            trip_stop_times = [(key, val) for key, val in trip if key not in special_keys and val]
            trip = {key: val for key, val in trip if key in special_keys}

            # Parse times from excel into proper format
            try:
                trip_stop_times = sorted([(key, (f'{val.total_seconds()//3600:02.0f}:{val.total_seconds()%3600//60:02.0f}:{val.total_seconds()%3600%60:02.0f}' if type(val)==datetime.timedelta else (str(val.day*24 + val.hour) if type(val)==datetime.datetime else val.strftime("%H")) + val.strftime(":%M:%S"))  ) for key, val in trip_stop_times], key=lambda x: x[-1])
            except:
                print(f'Error converting trip:\n{trip}')

            # Define trip_id and prevent duplicates
            trip_id = "-".join(str(item) for item in [service, *([trip.get("trip_short_name", "")] if trip.get("trip_short_name", "") else trip_stop_times[0])])
            while trip_id in trip_ids:
                trip_id += "+"
            trip_ids.append(trip_id)

            # Append trips.txt Entries
            gtfs_entries["trips"].append({
                "service_id": service,
                "trip_id": trip_id,
                "route_id": trip.get("route_id"),
                "direction_id": trip.get("direction_id"),
                "shape_id": trip.get("shape_id", ""),
                "trip_short_name": trip.get("trip_short_name", ""),
                "trip_headsign": trip.get("headsign", ""),
                "wheelchair_accessible": trip.get("wheelchair_accessible", ""),
                "bikes_allowed": trip.get("bikes_allowed", ""),
                "block_id": trip.get("block_id", ""),
            })

            # Append stop_time.txt Entries
            [gtfs_entries["stop_times"].append({
                "trip_id": trip_id,
                "arrival_time": val[1],
                "departure_time": val[1],
                "stop_id": str(val[0]),
                "stop_sequence": index,
                "pickup_type": "1" if index==len(trip_stop_times)-1 else "0",
                "drop_off_type": "1" if index==0 else "0"
            }) for index, val in enumerate(trip_stop_times)]

            # Append frequencies.txt Entries
            if trip.get("Then Every") and trip.get("Until"):
                gtfs_entries["frequencies"].append({
                    "trip_id": trip_id,
                    "start_time": trip_stop_times[0][-1],
                    "end_time": trip["Until"].strftime("%H:%M:%S"),
                    "headway_secs": (trip["Then Every"].hour*60*60 + trip["Then Every"].minute*60 + trip["Then Every"].second),
                    "exact_times": "1"
                })

    # ------------------------------------------------------------------------------
    # Write Schedule Data
    # ------------------------------------------------------------------------------

    for key in gtfs_entries:
        if gtfs_entries[key]:
            with open(f'{fp}/{key}.txt', "w") as file:
                writer = csv.DictWriter(file, fieldnames=list(gtfs_entries[key][0]))
                writer.writeheader()
                writer.writerows(gtfs_entries[key])


if __name__=="__main__":
    filepath = input("Enter the filepath for excel2gtfs conversion:\n> ") if len(sys.argv) < 2 else sys.argv[1]
    excel2gtfs(filepath)
